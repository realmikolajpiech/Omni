"""Tool definitions and execution for function calling with the main LLM."""
import json
import logging
import re
import threading

from src.core.stats_store import increment_tool

# ── Per-request trust-request collector ──────────────────────────────────────
# Each Flask request runs in its own thread, so threading.local() keeps
# the pending list isolated per request.
_tls = threading.local()


def _get_pending() -> list:
    if not hasattr(_tls, "trust_requests"):
        _tls.trust_requests = []
    return _tls.trust_requests


def flush_pending_trust_requests() -> list:
    """Return and clear all trust_request actions accumulated this request."""
    reqs = list(_get_pending())
    _tls.trust_requests = []
    return reqs


# ── Temporary trust boost (set by UI on "Allow once" for request_permission) ─
# Raised to the required level for a single AI turn; cleared in on_ai_response.
_trust_boost: int = 0


def set_trust_boost(level: int) -> None:
    """Temporarily elevate effective trust level for the current AI turn."""
    global _trust_boost
    _trust_boost = level


def clear_trust_boost() -> None:
    """Clear the temporary trust boost after a request completes."""
    global _trust_boost
    _trust_boost = 0


def get_effective_trust() -> int:
    """Return the higher of the user's configured trust level and any active boost."""
    import src.core.settings_store as _ss
    return max(_ss.get("trust_level", 1), _trust_boost)

# ── Tool Schemas ─────────────────────────────────────────────────────────────

TOOL_SCHEMAS = [
    {
        "type": "function",
        "function": {
            "name": "search_web",
            "description": (
                "Search the web for current information, news, recent events, facts, "
                "prices, weather, sports scores, or anything that may have changed "
                "recently or isn't in the model's training knowledge."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "The search query to look up on the web.",
                    }
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_files",
            "description": (
                "Semantically search user's local files and documents stored on the "
                "computer. Use when asked about personal notes, journals, code files, "
                "PDFs, spreadsheets, certificates, official documents, invoices, tax "
                "forms, contracts, resumes, or any content that might be saved locally. "
                "Also use when the user's query implies they want to FIND or LOCATE a "
                "document, even if they don't explicitly say 'find my file'. Works with "
                "queries in any language."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "What to search for in local files.",
                    }
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "calculate",
            "description": (
                "Evaluate a mathematical expression precisely. Use for arithmetic, "
                "algebra, percentages, and any numerical computation where accuracy matters."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "expression": {
                        "type": "string",
                        "description": "The mathematical expression to evaluate, e.g. '(100 * 1.23) / 4 + 7'.",
                    }
                },
                "required": ["expression"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_images",
            "description": (
                "Search user's local image library by visual description or content. "
                "Use when asked about photos or images stored on the computer."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Description of the image or photo to find.",
                    }
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "memory_recall",
            "description": (
                "Retrieve facts, preferences and personal details about this user from long-term memory. "
                "Use proactively when the answer could depend on something the user has mentioned before — "
                "their job, habits, goals, relationships, preferences, or anything personal. "
                "Also use when the user asks 'do you remember…' or 'what do you know about…'."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "What to look up in memory, e.g. 'diet preferences', 'job', 'workout routine'.",
                    }
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "memory_save",
            "description": (
                "Permanently save a new fact or preference about the user to long-term memory. "
                "MUST be called IMMEDIATELY when the user shares a name, preference, correction, or personal detail. "
                "Do not just acknowledge — save it."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "fact": {
                        "type": "string",
                        "description": "A concise, factual statement to remember, e.g. 'User is vegetarian' or 'User's girlfriend is called Ania'.",
                    }
                },
                "required": ["fact"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "run_terminal",
            "description": (
                "Execute a shell command on the user's macOS system and return its output. "
                "Use for any system task: reading battery/CPU/disk/RAM info, changing settings via "
                "defaults write or osascript, running scripts, managing files, getting uptime, etc. "
                "NEVER tell the user to open Terminal manually — just call this tool."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "command": {
                        "type": "string",
                        "description": "The shell command to execute, e.g. 'defaults write com.apple.dock autohide -bool true && killall Dock'.",
                    },
                    "description": {
                        "type": "string",
                        "description": "Brief human-readable description of what this command does, e.g. 'Enable Dock autohide'.",
                    },
                },
                "required": ["command"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "install_app",
            "description": (
                "Install an application on the user's macOS using Homebrew. "
                "Use when the user asks to install, download, or get any app or CLI tool. "
                "Tries --cask first (GUI apps like firefox, vlc, discord), falls back to formula (CLI tools like git, ffmpeg)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {
                        "type": "string",
                        "description": "The Homebrew package name, e.g. 'firefox', 'vlc', 'discord', 'git'.",
                    },
                },
                "required": ["name"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "uninstall_app",
            "description": (
                "Uninstall/remove an application from the user's macOS using Homebrew. "
                "Use when the user asks to uninstall, remove, or delete an app."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {
                        "type": "string",
                        "description": "The Homebrew package or cask name to remove, e.g. 'firefox', 'vlc', 'discord'.",
                    },
                },
                "required": ["name"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_calendar_events",
            "description": "Get upcoming calendar events from the macOS Calendar app.",
            "parameters": {
                "type": "object",
                "properties": {
                    "days": {
                        "type": "integer",
                        "description": "Number of days to look ahead (default 3).",
                    }
                },
                "required": [],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "create_calendar_event",
            "description": "Create a new event in the macOS Calendar app.",
            "parameters": {
                "type": "object",
                "properties": {
                    "title": {"type": "string", "description": "Title of the event."},
                    "start_iso": {"type": "string", "description": "Start time in 'YYYY-MM-DD HH:MM:SS' format."},
                    "duration_minutes": {"type": "integer", "description": "Duration in minutes (default 60)."},
                    "description": {"type": "string", "description": "Description or notes for the event."},
                },
                "required": ["title", "start_iso"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_unread_emails",
            "description": "Get recent unread emails from macOS Mail app.",
            "parameters": {
                "type": "object",
                "properties": {
                    "limit": {
                        "type": "integer",
                        "description": "Max number of emails to retrieve (default 5).",
                    }
                },
                "required": [],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "send_email",
            "description": "Compose an email for user review. A compose widget is shown with editable To/Subject/Body fields. Call this even if you don't know the recipient's email address — the user can fill it in.",
            "parameters": {
                "type": "object",
                "properties": {
                    "to": {"type": "string", "description": "Recipient email address. Leave empty string if unknown — user will fill it in."},
                    "subject": {"type": "string", "description": "Subject line."},
                    "body": {"type": "string", "description": "Email body content."},
                },
                "required": ["subject", "body"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "memory_delete",
            "description": (
                "Delete or forget a specific memory about the user. "
                "Use when the user explicitly asks you to forget something, or when you learn that "
                "a previously stored fact is now outdated or incorrect."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Describe what to forget, e.g. 'my old job at Google' or 'that I was vegetarian'.",
                    }
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "organize_folder",
            "description": (
                "Organize files in a folder into smart subfolders. "
                "Use when the user asks to 'cleanup', 'organize', or 'tidy up' a folder. "
                "Smart mode creates subcategories (e.g. Images/Photos, Images/Screenshots, Code/Python, Documents/PDFs, Documents/Spreadsheets) "
                "and automatically merges small groups to avoid clutter. "
                "Supports 100+ file types including design files, 3D models, datasets, and more."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "The directory path to organize, e.g. '~/Desktop' or '~/Downloads'.",
                    },
                    "strategy": {
                        "type": "string",
                        "enum": ["smart", "type", "date"],
                        "description": "Strategy: 'smart' (default, intelligent subcategories that auto-merge small groups), 'type' (broad categories only), 'date' (Year/Month folders).",
                    }
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "create_file",
            "description": (
                "Create a new file with specified content on the user's Mac. "
                "Use when asked to 'create', 'write', 'save', or 'make' a file (txt, md, csv, etc.). "
                "Defaults to ~/Desktop if no folder is given. "
                "Always prefer this over run_terminal for file creation."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "filename": {
                        "type": "string",
                        "description": "The filename with extension, e.g. 'receipt.txt', 'notes.md', 'data.csv'.",
                    },
                    "content": {
                        "type": "string",
                        "description": "The text content to write into the file.",
                    },
                    "folder": {
                        "type": "string",
                        "description": "Directory where the file should be created, e.g. '~/Desktop', '~/Documents'. Defaults to ~/Desktop.",
                    },
                },
                "required": ["filename", "content"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "edit_file",
            "description": (
                "Edit an existing file by replacing a specific snippet of text with new text. "
                "Use when the user asks to 'edit', 'update', 'modify', 'change', or 'fix' content in a file. "
                "First use find_file to locate the file, then run_terminal with 'cat' to read its content, "
                "then call this tool with the exact old text and the new replacement text. "
                "Always prefer this over run_terminal for file edits."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Absolute path to the file to edit, e.g. '/Users/oskar/Desktop/notes.txt'.",
                    },
                    "old_text": {
                        "type": "string",
                        "description": "The exact text snippet in the file to find and replace. Must match the file content exactly.",
                    },
                    "new_text": {
                        "type": "string",
                        "description": "The replacement text to insert in place of old_text.",
                    },
                },
                "required": ["path", "old_text", "new_text"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compress",
            "description": (
                "Compress files or folders into a ZIP archive. "
                "Use when the user asks to 'compress', 'zip', 'archive', or 'bundle' files or folders. "
                "Can compress a single file, multiple files, or an entire folder."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "paths": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "List of absolute file or folder paths to compress, e.g. ['/Users/oskar/Desktop/report.pdf', '/Users/oskar/Desktop/photos'].",
                    },
                    "output": {
                        "type": "string",
                        "description": "Optional output ZIP path. If omitted, creates the archive next to the first input path with a .zip extension.",
                    },
                },
                "required": ["paths"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "convert_file",
            "description": (
                "Convert a file to a different format. "
                "Supports: images (PNG, JPG, WEBP, BMP, TIFF, ICO, GIF), "
                "documents (PDF, DOCX, TXT, HTML, MD), "
                "audio (MP3, WAV, OGG, FLAC, M4A), "
                "video (MP4, MOV, AVI, MKV, WEBM, GIF). "
                "Use when user asks to convert, export, or change file format."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "input_path": {
                        "type": "string",
                        "description": "Path to the source file.",
                    },
                    "output_format": {
                        "type": "string",
                        "description": "Target format extension without dot (e.g. 'png', 'pdf', 'mp3', 'mp4').",
                    },
                    "output_path": {
                        "type": "string",
                        "description": "Optional custom output path. Defaults to same directory with new extension.",
                    },
                },
                "required": ["input_path", "output_format"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "read_file",
            "description": (
                "Read the content of any file — plain text, PDF, DOCX, XLSX, CSV, PPTX, RTF, code, etc. "
                "Use this whenever the user asks you to read, summarise, analyse, or answer questions about a specific file. "
                "Pass the exact path returned by find_file or search_files. "
                "Returns the extracted text content (truncated to 12 000 chars if very long)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Absolute path to the file, e.g. '/Users/mikolaj/Downloads/report.pdf'.",
                    },
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "find_file",
            "description": (
                "Find files or folders on the user's Mac by name or partial name. "
                "Returns exact file paths. Use this BEFORE deleting, moving, or opening a file "
                "when you need to locate its precise path. Also use when the user asks where a "
                "file is located or wants to confirm it exists. "
                "Prefer this over search_files when you need actionable paths, not content."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {
                        "type": "string",
                        "description": "The filename or partial name to search for, e.g. 'report.pdf', 'notes', 'project'.",
                    },
                    "folder": {
                        "type": "string",
                        "description": "Optional folder to limit the search to, e.g. '~/Desktop', '~/Documents'. Omit to search everywhere.",
                    },
                    "include_dirs": {
                        "type": "boolean",
                        "description": "Whether to include directories in results (default true).",
                    },
                },
                "required": ["name"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "set_reminder",
            "description": (
                "Create a reminder that fires as a macOS notification. "
                "Use for: 'remind me in 1 hour to call Jessy', 'every 10 minutes check my email'. "
                "For agentic/conditional reminders (e.g. 'notify me when Oskar emails me'), "
                "set a recurring interval and a query that the AI will run each time — "
                "include STOP_REMINDER in the query instructions so the AI knows when to stop."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "label": {
                        "type": "string",
                        "description": "Short human-readable title shown in the notification.",
                    },
                    "fire_at_iso": {
                        "type": "string",
                        "description": "ISO 8601 datetime for when the reminder should first fire, e.g. '2026-03-13T15:30:00'.",
                    },
                    "interval_seconds": {
                        "type": "integer",
                        "description": "If > 0, the reminder repeats every this many seconds after firing. Omit or 0 for one-shot.",
                    },
                    "query": {
                        "type": "string",
                        "description": (
                            "Optional AI query to run each time the reminder fires. "
                            "The result is shown in the notification. "
                            "For conditional reminders, instruct the AI to include STOP_REMINDER "
                            "in its response when the condition is met so the reminder stops."
                        ),
                    },
                },
                "required": ["label", "fire_at_iso"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "list_reminders",
            "description": "List all pending reminders the user has set.",
            "parameters": {
                "type": "object",
                "properties": {},
                "required": [],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "delete_reminder",
            "description": "Cancel and delete a pending reminder. Describe which reminder to delete in plain language (e.g. 'email reminder', 'Oskar call', 'bigos') — no need to look up IDs first.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Natural language description of the reminder to cancel, e.g. 'email summary', 'call Oskar', 'bigos'.",
                    }
                },
                "required": ["query"],
            },
        },
    },
    # ── Context Engine tools ──────────────────────────────────────────────
    {
        "type": "function",
        "function": {
            "name": "get_context",
            "description": (
                "Get the user's current work context — which app they're using, "
                "recent files they've worked on, related entities (people, events, emails). "
                "Use this to understand what the user is currently doing before answering "
                "questions about their work."
            ),
            "parameters": {
                "type": "object",
                "properties": {},
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_work_sessions",
            "description": (
                "List the user's recent work sessions with summaries. Each session "
                "includes which apps and files were used, duration, and a summary. "
                "Use this when the user asks what they were working on, or wants to "
                "resume a previous task."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "limit": {
                        "type": "integer",
                        "description": "Number of sessions to return (default 5).",
                    }
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "resume_session",
            "description": (
                "Resume a previous work session by reopening all files and apps "
                "from that session. The user should specify which session to resume "
                "(e.g., 'the one where I was working on the payment module')."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "session_id": {
                        "type": "string",
                        "description": "The ID of the session to resume.",
                    }
                },
                "required": ["session_id"],
            },
        },
    },
]


# ── Tool execution ────────────────────────────────────────────────────────────

def execute_tool(name: str, arguments: dict) -> str:
    """Dispatch a tool call by name and return the result as a plain string."""
    increment_tool(name)
    try:
        if name == "search_web":
            return _tool_search_web(arguments.get("query", ""))
        elif name == "search_files":
            return _tool_search_files(arguments.get("query", ""))
        elif name == "calculate":
            return _tool_calculate(arguments.get("expression", ""))
        elif name == "search_images":
            return _tool_search_images(arguments.get("query", ""))
        elif name == "memory_recall":
            return _tool_memory_recall(arguments.get("query", ""))
        elif name == "memory_save":
            return _tool_memory_save(arguments.get("fact", ""))
        elif name == "memory_delete":
            return _tool_memory_delete(arguments.get("query", ""))
        elif name == "run_terminal":
            return _tool_run_terminal(arguments.get("command", ""), arguments.get("description", ""))
        elif name == "install_app":
            return _tool_install_app(arguments.get("name", ""))
        elif name == "uninstall_app":
            return _tool_uninstall_app(arguments.get("name", ""))
        elif name == "get_calendar_events":
            return _tool_get_calendar_events(arguments.get("days", 3))
        elif name == "create_calendar_event":
            return _tool_create_calendar_event(
                arguments.get("title", ""),
                arguments.get("start_iso", ""),
                arguments.get("duration_minutes", 60),
                arguments.get("description", "")
            )
        elif name == "get_unread_emails":
            return _tool_get_unread_emails(arguments.get("limit", 5))
        elif name == "send_email":
            return _tool_send_email(
                arguments.get("to", ""),
                arguments.get("subject", ""),
                arguments.get("body", "")
            )
        elif name == "organize_folder":
            return _tool_organize_folder(arguments.get("path", ""), arguments.get("strategy", "smart"))
        elif name == "create_file":
            return _tool_create_file(
                arguments.get("filename", ""),
                arguments.get("content", ""),
                arguments.get("folder", ""),
            )
        elif name == "compress":
            return _tool_compress(
                arguments.get("paths", []),
                arguments.get("output", ""),
            )
        elif name == "edit_file":
            return _tool_edit_file(
                arguments.get("path", ""),
                arguments.get("old_text", ""),
                arguments.get("new_text", ""),
            )
        elif name == "read_file":
            return _tool_read_file(arguments.get("path", ""))
        elif name == "find_file":
            return _tool_find_file(
                arguments.get("name", ""),
                arguments.get("folder", ""),
                arguments.get("include_dirs", True),
            )
        elif name == "convert_file":
            return _tool_convert_file(
                arguments.get("input_path", ""),
                arguments.get("output_format", ""),
                arguments.get("output_path", ""),
            )
        elif name == "set_reminder":
            return _tool_set_reminder(
                arguments.get("label", ""),
                arguments.get("fire_at_iso", ""),
                arguments.get("interval_seconds", 0),
                arguments.get("query", ""),
            )
        elif name == "list_reminders":
            return _tool_list_reminders()
        elif name == "delete_reminder":
            return _tool_delete_reminder(arguments.get("query", ""))
        elif name == "get_context":
            return _tool_get_context()
        elif name == "get_work_sessions":
            return _tool_get_work_sessions(arguments.get("limit", 5))
        elif name == "resume_session":
            return _tool_resume_session(arguments.get("session_id", ""))
        else:
            return json.dumps({"error": f"Unknown tool: {name}"})
    except Exception as e:
        logging.error(f"[tools] Execution error in '{name}': {e}")
        return json.dumps({"error": str(e)})


# ── Draft tool names (these return proposals, not results) ────────────────────
DRAFT_TOOLS = {
    "set_reminder", "create_calendar_event", "create_file",
    "edit_file", "compress", "convert_file", "organize_folder",
    "run_terminal",
}


def execute_tool_draft(name: str, arguments: dict) -> str:
    """Execute a previously deferred tool draft. Called from the UI after user clicks."""
    import os
    try:
        if name == "set_reminder":
            from src.services.reminders.reminder_service import add_reminder
            from datetime import datetime
            fire_at = datetime.fromisoformat(arguments["fire_at_iso"].strip()).timestamp()
            rid = add_reminder(
                arguments["label"].strip(),
                fire_at,
                int(arguments.get("interval_seconds", 0)),
                arguments.get("query", "").strip(),
            )
            return f"ok (id:{rid})"

        elif name == "create_calendar_event":
            from src.services.system.productivity import create_calendar_event
            return create_calendar_event(
                arguments["title"].strip(),
                arguments["start_iso"].strip(),
                int(arguments.get("duration_minutes", 60)),
                arguments.get("description", ""),
            )

        elif name == "create_file":
            folder = os.path.expanduser(arguments.get("folder", "").strip() or "~/Desktop")
            os.makedirs(folder, exist_ok=True)
            file_path = os.path.join(folder, arguments["filename"].strip())
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(arguments.get("content", ""))
            return f"Created: {file_path}"

        elif name == "edit_file":
            path = os.path.expanduser(arguments["path"].strip())
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
            new_content = content.replace(arguments["old_text"], arguments["new_text"], 1)
            with open(path, "w", encoding="utf-8") as f:
                f.write(new_content)
            return f"Edited: {path}"

        elif name == "compress":
            import zipfile
            paths = arguments.get("paths", [])
            resolved = [os.path.expanduser(p.strip()) for p in paths]
            output = arguments.get("output", "")
            if output:
                zip_path = os.path.expanduser(output.strip())
            else:
                zip_path = resolved[0].rstrip("/") + ".zip"
            os.makedirs(os.path.dirname(zip_path) or ".", exist_ok=True)
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                for item in resolved:
                    if os.path.isfile(item):
                        zf.write(item, os.path.basename(item))
                    elif os.path.isdir(item):
                        base_dir = os.path.dirname(item)
                        for root, dirs, files in os.walk(item):
                            for fl in files:
                                full = os.path.join(root, fl)
                                zf.write(full, os.path.relpath(full, base_dir))
            size = os.path.getsize(zip_path)
            if size < 1024:
                size_str = f"{size} bytes"
            elif size < 1024 * 1024:
                size_str = f"{size / 1024:.1f} KB"
            else:
                size_str = f"{size / (1024 * 1024):.1f} MB"
            return f"Created: {zip_path} ({size_str})"

        elif name == "convert_file":
            return _tool_convert_file_execute(
                arguments.get("input_path", ""),
                arguments.get("output_format", ""),
                arguments.get("output_path", ""),
            )

        elif name == "organize_folder":
            from src.services.system.files import organize_folder
            return organize_folder(
                arguments.get("path", "").strip(),
                arguments.get("strategy", "smart"),
            )

        elif name == "run_terminal":
            import subprocess
            cmd = arguments.get("command", "").strip()
            if not cmd:
                return "Error: no command provided"
            result = subprocess.run(
                cmd, shell=True, capture_output=True, text=True, timeout=30
            )
            output = (result.stdout or "").strip()
            err = (result.stderr or "").strip()
            if result.returncode == 0:
                return f"Done: {output}" if output else "Done"
            else:
                return f"Error: {err or output or 'command failed'}"

        else:
            return f"Error: unknown draft tool '{name}'"
    except Exception as e:
        logging.error(f"[tools] Draft execution error in '{name}': {e}")
        return f"Error: {e}"


def _tool_search_web(query: str) -> str:
    from src.services.search.web_search import perform_web_search
    if not query.strip():
        return "Error: empty query."
    result = perform_web_search(query)
    return result or "No web results found."


def _tool_search_files(query: str) -> str:
    from src.services.search.local_search import perform_file_search
    if not query.strip():
        return "Error: empty query."
    result = perform_file_search(query)
    return result or "No local files found matching that query."


def _tool_calculate(expression: str) -> str:
    from src.services.llm.chat import perform_calculation
    if not expression.strip():
        return "Error: empty expression."
    return perform_calculation(expression)


def _tool_search_images(query: str) -> str:
    from src.services.search.image_search import perform_image_search_with_fallback
    if not query.strip():
        return "Error: empty query."
    result = perform_image_search_with_fallback(query)
    return result or "No matching images found."


def _tool_memory_recall(query: str) -> str:
    from src.services.memory.memvid_store import get_user_memory
    if not query.strip():
        return "Error: empty query."
    result = get_user_memory(query)
    return result or "No memories found for that query."


def _tool_memory_save(fact: str) -> str:
    from src.services.memory.memvid_store import remember_fact
    fact = fact.strip()
    if not fact:
        return "Error: empty fact."
    ok = remember_fact(fact)
    return f"Saved: {fact}" if ok else "Failed to save memory."


def _tool_memory_delete(query: str) -> str:
    from src.services.memory.memvid_store import delete_memory
    if not query.strip():
        return "Error: empty query."
    ok = delete_memory(query)
    return f"Deleted memories matching: {query}" if ok else "No matching memories found to delete."


# ── Terminal trust classification ─────────────────────────────────────────────

# Patterns that require trust level 3 (destructive / privileged)
_TERMINAL_L3 = [
    r'\brm\s',           # delete files
    r'\bsudo\b',         # privilege escalation
    r'\bchmod\b',        # permission change
    r'\bchown\b',        # ownership change
    r'\bdd\b',           # raw disk operations
    r'\bmkfs\b',         # format filesystem
    r'\bfdisk\b',        # disk partitioning
    r'diskutil\s+(erase|format|partition|zeroDisk|secureErase)',
    r'\bbrew\s+install\b',
    r'\bbrew\s+uninstall\b',
    r'\bnpm\s+install\b',
    r'\bpip\d?\s+install\b',
    r'\bapt(-get)?\s+install\b',
    r'\byum\s+install\b',
    r'\bdnf\s+install\b',
]

# Patterns that require trust level 2 (write to filesystem or system state)
_TERMINAL_L2 = [
    # ── File system writes ────────────────────────────────────────────────────
    r'\btouch\b',            # create / update file timestamp
    r'\bmkdir\b',            # create directory
    r'\bcp\b',               # copy file
    r'\bmv\b',               # move / rename file
    r'\btee\b',              # write to file while piping
    r'\bwget\b',             # download file to disk
    r'\bcurl\b.*\s-[a-z]*o', # curl saving output (-o / -O)
    r'(?<![=<>!])>{1,2}(?![>=])',  # shell redirect  >  or  >>
    # ── System state modifications ────────────────────────────────────────────
    r'\bdefaults\s+write\b',
    r'\bdefaults\s+delete\b',
    r'\bkillall\b',
    r'\bkill\s+(-\d+\s+)?\d+',
    r'\bosascript\b',
    r'\blaunchctl\b',
    r'\bpmset\b',
    r'\bnetworksetup\b',
    r'\bscutil\s+--set\b',
    r'\bsystemsetup\b',
]


def _terminal_required_trust_level(command: str) -> int:
    """Return the minimum trust level needed to run this terminal command.

    1 — read-only            (ioreg, df, system_profiler, sw_vers, cat, ls, …)
    2 — filesystem / system  (touch, mkdir, cp, mv, defaults write, killall, …)
    3 — destructive/privileged (rm, sudo, brew install, …)
    """
    lower = command.lower()
    for pat in _TERMINAL_L3:
        if re.search(pat, lower):
            return 3
    for pat in _TERMINAL_L2:
        if re.search(pat, lower):
            return 2
    return 1


def _tool_run_terminal(command: str, description: str = "") -> str:
    import subprocess

    command = command.strip()
    if not command:
        return "Error: empty command."

    required = _terminal_required_trust_level(command)
    current  = get_effective_trust()

    if current < required:
        # Queue a trust_request action so the UI can show a permission popup
        _get_pending().append({
            "type":           "trust_request",
            "required_level": required,
            "command":        command,
            "description":    description or command,
        })
        level_names = {1: "Assistant", 2: "Automation", 3: "Full Control"}
        return (
            f"[Permission required] '{level_names[required]}' trust is needed for this command. "
            f"The user is being prompted for one-time permission."
        )

    try:
        proc = subprocess.run(
            command, shell=True, capture_output=True, text=True, timeout=15
        )
        stdout = proc.stdout.strip()
        stderr = proc.stderr.strip()
        output = stdout
        if stderr:
            output += ("\n" if output else "") + f"STDERR: {stderr}"
        if not output:
            output = f"Done (exit code {proc.returncode})"
        return output
    except subprocess.TimeoutExpired:
        return "Error: command timed out after 15 seconds."
    except Exception as e:
        return f"Error: {e}"


def _tool_get_calendar_events(days: int) -> str:
    from src.services.system.productivity import get_calendar_events
    return get_calendar_events(days=int(days))


def _tool_create_calendar_event(title: str, start_iso: str, duration_minutes: int, description: str) -> str:
    title = title.strip()
    start_iso = start_iso.strip()
    if not title or not start_iso:
        return "Error: title and start_iso are required."
    # Don't create yet — UI will show a proposal widget
    return f"Calendar event draft prepared. Title: {title}, Start: {start_iso}, Duration: {duration_minutes} min"


def _tool_get_unread_emails(limit: int) -> str:
    from src.services.system.productivity import get_unread_emails
    return get_unread_emails(limit=int(limit))


def _tool_send_email(to: str, subject: str, body: str) -> str:
    to = to.strip()
    subject = subject.strip()
    body = body.strip()
    if not subject:
        return "Error: subject is required."
    # Don't send yet — the UI will show a compose widget for user review.
    return f"Email draft prepared for user review. To: {to or '(user will fill in)'}, Subject: {subject}"



def _find_brew() -> str | None:
    import os
    for path in ("/opt/homebrew/bin/brew", "/usr/local/bin/brew"):
        if os.path.exists(path):
            return path
    return None


def _tool_install_app(name: str) -> str:
    import subprocess, os
    name = name.strip()
    if not name:
        return "Error: empty app name."

    # Trust level 3 required for installing software
    current = get_effective_trust()
    if current < 3:
        _get_pending().append({
            "type":           "trust_request",
            "required_level": 3,
            "command":        f"brew install {name}",
            "description":    f"Install {name} via Homebrew",
        })
        return "[Permission required] 'Full Control' trust is needed to install software."

    from src.services.system.installer import generate_install_plan
    plan = generate_install_plan(name)
    if plan["method"] == "failed":
        return f"Error: {plan['description']}"

    commands = plan.get("commands", [])
    if not commands:
        return f"Error: no install command found for '{name}'."

    env = {**os.environ, "HOMEBREW_NO_AUTO_UPDATE": "1", "NONINTERACTIVE": "1",
           "PATH": "/opt/homebrew/bin:/usr/local/bin:" + os.environ.get("PATH", "")}
    try:
        result = subprocess.run(
            commands[0], shell=True, capture_output=True, text=True, timeout=300, env=env
        )
        output = (result.stdout + "\n" + result.stderr).strip()
        # Refresh app cache so the newly installed app is discoverable immediately
        try:
            import src.services.system.app_launcher as _al
            _al.APP_CACHE = None
        except Exception:
            pass
        return output[:1200] if output else f"Installed {name}"
    except subprocess.TimeoutExpired:
        return "Error: install timed out after 5 minutes."
    except Exception as e:
        return f"Error: {e}"


def _tool_uninstall_app(name: str) -> str:
    import subprocess, os
    name = name.strip()
    if not name:
        return "Error: empty app name."

    # Trust level 3 required for uninstalling software
    current = get_effective_trust()
    if current < 3:
        _get_pending().append({
            "type":           "trust_request",
            "required_level": 3,
            "command":        f"brew uninstall {name}",
            "description":    f"Uninstall {name} via Homebrew",
        })
        return "[Permission required] 'Full Control' trust is needed to uninstall software."

    from src.services.system.installer import generate_uninstall_plan
    plan = generate_uninstall_plan(name)
    if plan["method"] == "failed":
        return f"Error: {plan['description']}"
    if plan["method"] == "not_installed":
        return plan["description"]

    commands = plan.get("commands", [])
    if not commands:
        return f"Error: no uninstall command found for '{name}'."

    env = {**os.environ, "HOMEBREW_NO_AUTO_UPDATE": "1", "NONINTERACTIVE": "1",
           "PATH": "/opt/homebrew/bin:/usr/local/bin:" + os.environ.get("PATH", "")}
    try:
        result = subprocess.run(
            commands[0], shell=True, capture_output=True, text=True, timeout=120, env=env
        )
        output = (result.stdout + "\n" + result.stderr).strip()
        # Refresh app cache so the uninstalled app is no longer discoverable
        try:
            import src.services.system.app_launcher as _al
            _al.APP_CACHE = None
        except Exception:
            pass
        return output[:1200] if output else f"Uninstalled {name}"
    except subprocess.TimeoutExpired:
        return "Error: uninstall timed out."
    except Exception as e:
        return f"Error: {e}"

def _tool_organize_folder(path: str, strategy: str) -> str:
    path = path.strip()
    if not path:
        return "Error: empty path."
    # Don't organize yet — UI will show a proposal widget
    return f"Organize draft prepared. Path: {path}, Strategy: {strategy}"


def _tool_create_file(filename: str, content: str, folder: str = "") -> str:
    """Create a file with given content — returns draft for UI proposal widget."""
    filename = filename.strip()
    if not filename:
        return "Error: empty filename."
    # Don't create yet — UI will show a proposal widget
    folder = folder.strip() if folder and folder.strip() else "~/Desktop"
    return f"File creation draft prepared. Filename: {filename}, Folder: {folder}"


def _tool_compress(paths: list, output: str = "") -> str:
    """Compress files/folders — returns draft for UI proposal widget."""
    import os
    if not paths:
        return "Error: no paths provided."
    # Validate paths exist before proposing
    for p in paths:
        expanded = os.path.expanduser(p.strip())
        if not os.path.exists(expanded):
            return f"Error: path not found: {expanded}"
    # Don't compress yet — UI will show a proposal widget
    return f"Compress draft prepared. {len(paths)} item(s)"


def _tool_edit_file(path: str, old_text: str, new_text: str) -> str:
    """Edit a file — returns draft for UI proposal widget."""
    import os
    path = path.strip()
    if not path:
        return "Error: empty path."
    if not old_text:
        return "Error: old_text is required."
    # Validate that the file and old_text exist before proposing
    expanded = os.path.expanduser(path)
    if not os.path.isfile(expanded):
        return f"Error: file not found: {expanded}"
    try:
        with open(expanded, "r", encoding="utf-8") as f:
            content = f.read()
    except Exception as e:
        return f"Error reading file: {e}"
    if old_text not in content:
        return f"Error: old_text not found in {path}. Make sure it matches the file content exactly."
    # Don't edit yet — UI will show a proposal widget
    return f"File edit draft prepared. Path: {path}"


def _tool_read_file(path: str) -> str:
    """Read and return the text content of any supported file type."""
    import os
    path = os.path.expanduser(path.strip())
    if not path:
        return "Error: no path provided."
    if not os.path.exists(path):
        return f"Error: file not found: {path}"
    if os.path.isdir(path):
        return f"Error: '{path}' is a directory, not a file."
    if os.path.getsize(path) > 20 * 1024 * 1024:
        return "Error: file is too large to read (>20 MB)."

    from src.services.search.utils import _CONTENT_READERS, strip_rtf
    _, ext = os.path.splitext(path)
    ext = ext.lower()

    try:
        reader = _CONTENT_READERS.get(ext)
        if reader:
            content = reader(path)
        elif ext == ".rtf":
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                content = strip_rtf(f.read())
        else:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read()
    except Exception as e:
        return f"Error reading file: {e}"

    if not content or not content.strip():
        return "File is empty or its content could not be extracted."

    max_chars = 12000
    truncated = len(content) > max_chars
    result = content[:max_chars]
    if truncated:
        result += f"\n\n[... truncated — showing first {max_chars} of {len(content)} characters]"
    return result


def _tool_find_file(name: str, folder: str = "", include_dirs: bool = True) -> str:
    """Find files/folders by name using Spotlight (mdfind) with a find fallback."""
    import subprocess, os, sys

    name = name.strip()
    if not name:
        return "Error: empty name."

    folder = folder.strip()
    if folder:
        folder = os.path.expanduser(folder)

    results: list[str] = []

    # ── macOS: use mdfind (Spotlight) — instant and index-backed ─────────────
    if sys.platform == "darwin":
        try:
            cmd = ["mdfind", "-name", name]
            if folder:
                cmd += ["-onlyin", folder]
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=8)
            if proc.returncode == 0:
                for line in proc.stdout.strip().splitlines():
                    line = line.strip()
                    if not line or not os.path.exists(line):
                        continue
                    if not include_dirs and os.path.isdir(line):
                        continue
                    # Filter out noisy system paths
                    if any(seg in line for seg in (
                        "/Library/Caches/", "/.Spotlight-", "/.Trashes",
                        "/System/", "/private/var/", "/.vol/",
                    )):
                        continue
                    results.append(line)
        except Exception as e:
            logging.warning(f"[tool:find_file] mdfind failed: {e}")

    # ── Fallback: plain `find` (slower, used on non-mac or mdfind failure) ───
    if not results:
        search_root = folder or os.path.expanduser("~")
        try:
            cmd = ["find", search_root, "-maxdepth", "6", "-iname", f"*{name}*"]
            if not include_dirs:
                cmd += ["-type", "f"]
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
            for line in proc.stdout.strip().splitlines():
                line = line.strip()
                if line and os.path.exists(line):
                    results.append(line)
        except Exception as e:
            logging.warning(f"[tool:find_file] find fallback failed: {e}")

    if not results:
        return f"No files or folders named '{name}' found."

    # De-duplicate and cap output
    seen: set[str] = set()
    unique: list[str] = []
    for r in results:
        if r not in seen:
            seen.add(r)
            unique.append(r)

    unique = unique[:20]  # cap at 20 results

    lines = []
    for p in unique:
        kind = "dir" if os.path.isdir(p) else "file"
        try:
            size = os.path.getsize(p) if kind == "file" else 0
            size_str = f"  ({size:,} bytes)" if size else ""
        except OSError:
            size_str = ""
        lines.append(f"[{kind}] {p}{size_str}")

    return "\n".join(lines)


# ── File format sets ─────────────────────────────────────────────────────────
_IMAGE_EXTS = {"png", "jpg", "jpeg", "webp", "bmp", "tiff", "tif", "ico", "gif"}
_DOC_EXTS = {"pdf", "docx", "txt", "html", "htm", "md", "rtf", "csv", "xlsx"}
_AUDIO_EXTS = {"mp3", "wav", "ogg", "flac", "m4a", "aac", "wma"}
_VIDEO_EXTS = {"mp4", "mov", "avi", "mkv", "webm", "gif"}


def _format_size(size: int) -> str:
    if size < 1024:
        return f"{size} bytes"
    elif size < 1024 * 1024:
        return f"{size / 1024:.1f} KB"
    else:
        return f"{size / (1024 * 1024):.1f} MB"


def _tool_convert_file(input_path: str, output_format: str, output_path: str = "") -> str:
    """Convert a file — returns draft for UI proposal widget."""
    import os
    input_path = input_path.strip()
    output_format = output_format.strip().lower().lstrip(".")
    if not input_path:
        return "Error: empty input path."
    if not output_format:
        return "Error: empty output format."
    expanded = os.path.expanduser(input_path)
    if not os.path.isfile(expanded):
        return f"Error: file not found: {expanded}"
    # Don't convert yet — UI will show a proposal widget
    return f"Convert draft prepared. {os.path.basename(input_path)} to .{output_format}"


def _tool_convert_file_execute(input_path: str, output_format: str, output_path: str = "") -> str:
    """Actually execute file conversion (called from UI after user confirms)."""
    import os
    import subprocess
    import shutil

    input_path = input_path.strip()
    output_format = output_format.strip().lower().lstrip(".")

    input_path = os.path.expanduser(input_path)
    if not os.path.isfile(input_path):
        return f"Error: file not found: {input_path}"

    src_ext = os.path.splitext(input_path)[1].lower().lstrip(".")

    # Normalize jpeg → jpg
    if src_ext == "jpeg":
        src_ext = "jpg"
    if output_format == "jpeg":
        output_format = "jpg"

    if src_ext == output_format:
        return f"Error: file is already in .{output_format} format."

    # Determine output path
    if output_path:
        out = os.path.expanduser(output_path.strip())
    else:
        base = os.path.splitext(input_path)[0]
        out = f"{base}.{output_format}"

    # Avoid overwriting
    if os.path.exists(out):
        name, ext = os.path.splitext(out)
        counter = 1
        while os.path.exists(out):
            out = f"{name}_{counter}{ext}"
            counter += 1

    os.makedirs(os.path.dirname(out) or ".", exist_ok=True)

    logging.info(f"[tool:convert_file] {input_path} ({src_ext}) → {out} ({output_format})")

    try:
        # ── Image conversions (Pillow) ────────────────────────────────────
        if src_ext in _IMAGE_EXTS and output_format in _IMAGE_EXTS:
            return _convert_image(input_path, out, output_format)

        # ── Image → PDF ───────────────────────────────────────────────────
        if src_ext in _IMAGE_EXTS and output_format == "pdf":
            return _convert_image_to_pdf(input_path, out)

        # ── PDF → Image ───────────────────────────────────────────────────
        if src_ext == "pdf" and output_format in _IMAGE_EXTS:
            return _convert_pdf_to_image(input_path, out, output_format)

        # ── Document conversions (macOS textutil) ─────────────────────────
        if src_ext in ("docx", "doc", "rtf", "html", "htm", "txt") and \
           output_format in ("pdf", "docx", "doc", "rtf", "html", "txt"):
            return _convert_doc_textutil(input_path, out, output_format)

        # ── Markdown → HTML ───────────────────────────────────────────────
        if src_ext == "md" and output_format == "html":
            return _convert_md_to_html(input_path, out)

        # ── CSV → XLSX or XLSX → CSV ─────────────────────────────────────
        if (src_ext == "csv" and output_format == "xlsx") or \
           (src_ext == "xlsx" and output_format == "csv"):
            return _convert_spreadsheet(input_path, out, src_ext, output_format)

        # ── Audio conversions (ffmpeg) ────────────────────────────────────
        if src_ext in _AUDIO_EXTS and output_format in _AUDIO_EXTS:
            return _convert_ffmpeg(input_path, out, "audio")

        # ── Video conversions (ffmpeg) ────────────────────────────────────
        if src_ext in _VIDEO_EXTS and output_format in _VIDEO_EXTS:
            return _convert_ffmpeg(input_path, out, "video")

        # ── Video → Audio (extract audio) ─────────────────────────────────
        if src_ext in _VIDEO_EXTS and output_format in _AUDIO_EXTS:
            return _convert_ffmpeg(input_path, out, "audio_extract")

        # ── Video → GIF ───────────────────────────────────────────────────
        if src_ext in _VIDEO_EXTS and output_format == "gif":
            return _convert_ffmpeg(input_path, out, "video_to_gif")

        return f"Error: unsupported conversion from .{src_ext} to .{output_format}"

    except Exception as e:
        logging.error(f"[tool:convert_file] {e}")
        return f"Error converting file: {e}"


def _convert_image(input_path: str, out: str, fmt: str) -> str:
    """Convert between image formats using Pillow."""
    import os
    from PIL import Image

    img = Image.open(input_path)

    # Handle transparency for formats that don't support it
    if fmt in ("jpg", "bmp", "ico") and img.mode in ("RGBA", "LA", "P"):
        bg = Image.new("RGB", img.size, (255, 255, 255))
        if img.mode == "P":
            img = img.convert("RGBA")
        bg.paste(img, mask=img.split()[-1] if img.mode == "RGBA" else None)
        img = bg

    save_fmt = {"jpg": "JPEG", "tif": "TIFF"}.get(fmt, fmt.upper())
    img.save(out, format=save_fmt)
    size = os.path.getsize(out)
    return f"Converted: {out} ({_format_size(size)})"


def _convert_image_to_pdf(input_path: str, out: str) -> str:
    """Convert image to PDF using Pillow."""
    import os
    from PIL import Image

    img = Image.open(input_path)
    if img.mode == "RGBA":
        bg = Image.new("RGB", img.size, (255, 255, 255))
        bg.paste(img, mask=img.split()[-1])
        img = bg
    elif img.mode != "RGB":
        img = img.convert("RGB")

    img.save(out, format="PDF")
    size = os.path.getsize(out)
    return f"Converted: {out} ({_format_size(size)})"


def _convert_pdf_to_image(input_path: str, out: str, fmt: str) -> str:
    """Convert PDF pages to images using PyMuPDF (fitz) or sips fallback."""
    import os
    import subprocess

    # Try PyMuPDF first
    try:
        import fitz
        doc = fitz.open(input_path)
        if len(doc) == 1:
            page = doc[0]
            pix = page.get_pixmap(dpi=200)
            pix.save(out)
            size = os.path.getsize(out)
            return f"Converted: {out} ({_format_size(size)})"
        else:
            # Multi-page: save each page
            base, ext = os.path.splitext(out)
            paths = []
            for i, page in enumerate(doc):
                pix = page.get_pixmap(dpi=200)
                page_out = f"{base}_page{i + 1}{ext}"
                pix.save(page_out)
                paths.append(page_out)
            return f"Converted {len(paths)} pages: {paths[0]} ... {paths[-1]}"
    except ImportError:
        pass

    # Fallback: macOS sips (only for single-page, first page via Preview)
    try:
        proc = subprocess.run(
            ["sips", "-s", "format", fmt, input_path, "--out", out],
            capture_output=True, text=True, timeout=30,
        )
        if proc.returncode == 0 and os.path.isfile(out):
            size = os.path.getsize(out)
            return f"Converted: {out} ({_format_size(size)})"
    except Exception:
        pass

    return "Error: PDF to image conversion requires PyMuPDF (pip install pymupdf). Install it for best results."


def _convert_doc_textutil(input_path: str, out: str, fmt: str) -> str:
    """Convert documents using macOS textutil."""
    import os
    import subprocess
    import sys

    if sys.platform != "darwin":
        return "Error: document conversion via textutil is only available on macOS."

    # textutil format names
    fmt_map = {
        "txt": "txt", "html": "html", "htm": "html",
        "rtf": "rtf", "docx": "docx", "doc": "doc",
        "pdf": "pdf",
    }
    tu_fmt = fmt_map.get(fmt)
    if not tu_fmt:
        return f"Error: textutil doesn't support .{fmt} output."

    # textutil can't do PDF directly — use it to make HTML, then wkhtmltopdf or cupsfilter
    if tu_fmt == "pdf":
        # Convert to HTML first, then to PDF via cupsfilter
        html_tmp = out + ".tmp.html"
        try:
            proc = subprocess.run(
                ["textutil", "-convert", "html", "-output", html_tmp, input_path],
                capture_output=True, text=True, timeout=30,
            )
            if proc.returncode != 0:
                return f"Error: textutil failed: {proc.stderr.strip()}"

            # Try cupsfilter for HTML→PDF
            proc2 = subprocess.run(
                ["cupsfilter", html_tmp],
                capture_output=True, timeout=30,
            )
            if proc2.returncode == 0 and proc2.stdout:
                with open(out, "wb") as f:
                    f.write(proc2.stdout)
                os.unlink(html_tmp)
                size = os.path.getsize(out)
                return f"Converted: {out} ({_format_size(size)})"

            # Fallback: just keep the HTML
            os.rename(html_tmp, out.replace(".pdf", ".html"))
            return f"Converted to HTML (PDF conversion needs wkhtmltopdf): {out.replace('.pdf', '.html')}"
        except Exception as e:
            if os.path.exists(html_tmp):
                os.unlink(html_tmp)
            return f"Error converting to PDF: {e}"

    proc = subprocess.run(
        ["textutil", "-convert", tu_fmt, "-output", out, input_path],
        capture_output=True, text=True, timeout=30,
    )
    if proc.returncode != 0:
        return f"Error: textutil failed: {proc.stderr.strip()}"

    if not os.path.isfile(out):
        return "Error: conversion produced no output."

    size = os.path.getsize(out)
    return f"Converted: {out} ({_format_size(size)})"


def _convert_md_to_html(input_path: str, out: str) -> str:
    """Convert Markdown to HTML."""
    import os

    with open(input_path, "r", encoding="utf-8") as f:
        md_text = f.read()

    try:
        import markdown
        html = markdown.markdown(md_text, extensions=["tables", "fenced_code"])
    except ImportError:
        # Simple fallback: wrap in <pre>
        import html as html_mod
        html = f"<html><body><pre>{html_mod.escape(md_text)}</pre></body></html>"

    with open(out, "w", encoding="utf-8") as f:
        f.write(html)

    size = os.path.getsize(out)
    return f"Converted: {out} ({_format_size(size)})"


def _convert_spreadsheet(input_path: str, out: str, src_ext: str, dst_ext: str) -> str:
    """Convert between CSV and XLSX."""
    import os
    import csv

    if src_ext == "csv" and dst_ext == "xlsx":
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            with open(input_path, "r", encoding="utf-8") as f:
                for row in csv.reader(f):
                    ws.append(row)
            wb.save(out)
            size = os.path.getsize(out)
            return f"Converted: {out} ({_format_size(size)})"
        except ImportError:
            return "Error: CSV to XLSX conversion requires openpyxl (pip install openpyxl)."

    elif src_ext == "xlsx" and dst_ext == "csv":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(input_path, read_only=True)
            ws = wb.active
            with open(out, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)
            size = os.path.getsize(out)
            return f"Converted: {out} ({_format_size(size)})"
        except ImportError:
            return "Error: XLSX to CSV conversion requires openpyxl (pip install openpyxl)."

    return f"Error: unsupported spreadsheet conversion."


def _convert_ffmpeg(input_path: str, out: str, mode: str) -> str:
    """Convert audio/video using ffmpeg."""
    import os
    import subprocess
    import shutil

    if not shutil.which("ffmpeg"):
        return "Error: ffmpeg not found. Install it with: brew install ffmpeg"

    if mode == "audio_extract":
        cmd = ["ffmpeg", "-i", input_path, "-vn", "-y", out]
    elif mode == "video_to_gif":
        cmd = [
            "ffmpeg", "-i", input_path,
            "-vf", "fps=15,scale=480:-1:flags=lanczos",
            "-y", out,
        ]
    else:
        cmd = ["ffmpeg", "-i", input_path, "-y", out]

    proc = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if proc.returncode != 0:
        err = proc.stderr.strip().split("\n")[-1] if proc.stderr else "unknown error"
        return f"Error: ffmpeg failed: {err}"

    if not os.path.isfile(out):
        return "Error: conversion produced no output."

    size = os.path.getsize(out)
    return f"Converted: {out} ({_format_size(size)})"


# ── Reminder tools ─────────────────────────────────────────────────────────────

def _tool_set_reminder(label: str, fire_at_iso: str, interval_seconds: int = 0, query: str = "") -> str:
    label = label.strip()
    fire_at_iso = fire_at_iso.strip()
    if not label or not fire_at_iso:
        return "Error: label and fire_at_iso are required."
    # Validate datetime but don't create yet — UI will show a proposal widget
    from datetime import datetime
    try:
        datetime.fromisoformat(fire_at_iso)
    except ValueError as e:
        return f"Error: invalid ISO datetime '{fire_at_iso}': {e}"
    return f"Reminder draft prepared. Label: {label}, At: {fire_at_iso}"


def _tool_list_reminders() -> str:
    from src.services.reminders.reminder_service import list_reminders
    from datetime import datetime
    reminders = list_reminders()
    if not reminders:
        return "No pending reminders."
    lines = []
    for r in reminders:
        when = datetime.fromtimestamp(r["fire_at"]).strftime("%Y-%m-%d %H:%M:%S")
        parts = [f"[{r['id']}] {r['label']!r} — fires at {when}"]
        if r.get("interval_seconds"):
            parts.append(f"repeats every {r['interval_seconds']}s")
        if r.get("query"):
            q = r["query"]
            if len(q) > 60:
                q = q[:57] + "…"
            parts.append(f"query: {q!r}")
        lines.append(", ".join(parts))
    return "\n".join(lines)


def _tool_delete_reminder(query: str) -> str:
    from src.services.reminders.reminder_service import delete_reminder, list_reminders
    query = query.strip().lower()
    if not query:
        return "Error: query is required."
    reminders = list_reminders()
    if not reminders:
        return "No pending reminders to delete."
    # Find best match: exact ID, then substring match on label
    match = next((r for r in reminders if r["id"] == query), None)
    if not match:
        match = next((r for r in reminders if query in r["label"].lower()), None)
    if not match:
        # Try matching any word from the query against the label
        words = query.split()
        match = next(
            (r for r in reminders if any(w in r["label"].lower() for w in words if len(w) > 2)),
            None
        )
    if not match:
        labels = ", ".join(f'"{r["label"]}"' for r in reminders)
        return f"No reminder matching '{query}'. Pending reminders: {labels}."
    delete_reminder(match["id"])
    return f"ok (cancelled '{match['label']}')"


# ── Context Engine tools ─────────────────────────────────────────────────────

def _tool_get_context() -> str:
    """Get the user's current work context."""
    try:
        from src.services.context.knowledge_graph import get_knowledge_graph
        from src.services.context.context_matcher import get_matcher
        kg = get_knowledge_graph()
        matcher = get_matcher()

        # Active entities (what the user is working on right now)
        active_ids = kg.get_active_entity_ids(window_seconds=300)
        active_entities = [kg.get_entity(eid) for eid in active_ids if eid]
        active_entities = [e for e in active_entities if e]

        # Most relevant entities (scored by context + recency)
        relevant = matcher.get_relevant_entities(limit=5)

        # Recent activity summary
        recent = kg.get_recent_activity(limit=10)

        parts = []
        if active_entities:
            parts.append("Currently active:")
            for e in active_entities[:5]:
                uri = e.get("uri", "")
                parts.append(f"  - {e['type']}: {e['name']}" + (f" ({uri})" if uri else ""))

        if relevant:
            parts.append("\nMost relevant entities:")
            for e in relevant[:5]:
                score = e.get("relevance_score", 0)
                parts.append(f"  - {e['type']}: {e['name']} (relevance: {score:.2f})")

        if recent:
            parts.append("\nRecent activity:")
            for a in recent[:5]:
                app = a.get("app_name", "")
                title = a.get("window_title", "")
                dur = a.get("duration_s", 0)
                parts.append(f"  - {app}: {title} ({int(dur)}s)")

        stats = kg.get_stats()
        parts.append(f"\nGraph stats: {stats['entities']} entities, {stats['relationships']} relationships")

        return "\n".join(parts) if parts else "No context data available yet."
    except Exception as e:
        return f"Context engine not available: {e}"


def _tool_get_work_sessions(limit: int = 5) -> str:
    """List recent work sessions."""
    try:
        from src.services.context.session_manager import get_session_manager
        mgr = get_session_manager()
        sessions = mgr.get_recent_sessions(limit=limit)

        if not sessions:
            return "No work sessions recorded yet."

        parts = []
        for s in sessions:
            import time as _time
            start = _time.strftime("%Y-%m-%d %H:%M", _time.localtime(s["start_time"]))
            end = _time.strftime("%H:%M", _time.localtime(s["end_time"]))
            duration_min = max(1, int((s["end_time"] - s["start_time"]) / 60))
            summary = s.get("summary", "No summary")
            resume = s.get("resume_state", {})
            files = [ap.get("path", "") for ap in resume.get("app_paths", [])]
            file_list = ", ".join(files[:3]) if files else "none"

            parts.append(
                f"Session {s['id']} ({start}–{end}, {duration_min} min):\n"
                f"  Summary: {summary}\n"
                f"  Files: {file_list}"
            )

        return "\n\n".join(parts)
    except Exception as e:
        return f"Session manager not available: {e}"


def _tool_resume_session(session_id: str) -> str:
    """Resume a work session by reopening its files/apps."""
    if not session_id:
        return "Error: session_id is required."
    try:
        from src.services.context.session_manager import get_session_manager, SessionManager
        mgr = get_session_manager()
        session = mgr.get_session(session_id)
        if not session:
            return f"Session '{session_id}' not found."
        return SessionManager.resume_session(session)
    except Exception as e:
        return f"Failed to resume session: {e}"
